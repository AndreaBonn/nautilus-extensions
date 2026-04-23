"""
excel_preview.py — Nautilus extension for Excel file preview
=============================================================
Adds "Anteprima Excel" to the right-click menu on .xlsx, .xls, .ods files.
Shows sheets, data, statistics, and metadata.

Installation:
    cp excel_preview.py ~/.local/share/nautilus-python/extensions/
    nautilus -q && nautilus

Dependencies:
    sudo apt install python3-openpyxl python3-pandas
"""

import logging
import os
import subprocess
import threading

import gi

gi.require_version("Nautilus", "4.0")
gi.require_version("Gtk", "4.0")
gi.require_version("GLib", "2.0")

from gi.repository import GLib, GObject, Gtk, Nautilus, Pango

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".ods"}

PREVIEW_ROWS = 100
MIN_COL_WIDTH = 80
MAX_COL_WIDTH = 300
WINDOW_W = 1150
WINDOW_H = 700

CSS = b"""
.xl-info-bar {
    background-color: #f6f8fa;
    border-bottom: 1px solid #e1e4e8;
    padding: 6px 12px;
}
.xl-stat-box {
    background-color: #ffffff;
    border: 1px solid #e1e4e8;
    border-radius: 6px;
    padding: 8px 14px;
    margin: 4px;
}
.xl-stat-title { font-size: 11px; color: #6a737d; }
.xl-stat-value { font-size: 15px; font-weight: bold; color: #24292e; }
.xl-sheet-tab  { font-size: 12px; }
.mono { font-family: monospace; font-size: 13px; }
"""


def fmt_size(size: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} TB"


# --------------------------------------------------------------------------- #
# Excel file reading
# --------------------------------------------------------------------------- #


def read_excel(path: str) -> dict:
    try:
        file_size = fmt_size(os.path.getsize(path))
    except OSError as e:
        logging.warning("Cannot stat Excel file %s: %s", path, e)
        file_size = "—"

    result = {
        "path": path,
        "error": None,
        "sheets": [],
        "metadata": {},
        "file_size": file_size,
    }

    try:
        # Import lazy — eseguito solo al primo click, poi in cache
        import openpyxl
        import pandas as pd

        # Una sola apertura del workbook per metadati + lista fogli
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        props = wb.properties
        result["metadata"] = {
            "Titolo": props.title or "—",
            "Autore": props.creator or "—",
            "Creato": str(props.created)[:19].replace("T", " ") if props.created else "—",
            "Modificato": str(props.modified)[:19].replace("T", " ") if props.modified else "—",
            "Fogli": str(len(wb.sheetnames)),
        }

        # Conta righe per ogni foglio direttamente da openpyxl (read_only, veloce)
        sheet_row_counts = {}
        for ws in wb.worksheets:
            sheet_row_counts[ws.title] = ws.max_row - 1  # -1 per l'header

        sheet_names = wb.sheetnames
        wb.close()

        # Leggi tutti i fogli in una sola chiamata pandas (molto più veloce)
        all_sheets = pd.read_excel(
            path,
            sheet_name=None,  # None = tutti i fogli
            nrows=PREVIEW_ROWS + 1,
            engine="openpyxl",
        )

        for sheet_name in sheet_names:
            try:
                df = all_sheets.get(sheet_name)
                if df is None:
                    raise ValueError("Foglio non trovato")

                truncated = len(df) > PREVIEW_ROWS
                df_preview = df.head(PREVIEW_ROWS)
                total_rows = sheet_row_counts.get(sheet_name, len(df_preview))

                numeric_cols = list(df_preview.select_dtypes(include="number").columns)
                null_counts = df_preview.isnull().sum().to_dict()

                sheet_data = {
                    "name": sheet_name,
                    "total_rows": total_rows,
                    "total_cols": len(df.columns),
                    "truncated": truncated,
                    "headers": list(df_preview.columns.astype(str)),
                    "rows": df_preview.astype(str).values.tolist(),
                    "numeric_cols": numeric_cols,
                    "null_counts": null_counts,
                    "dtypes": {c: str(df_preview[c].dtype) for c in df_preview.columns},
                    "describe": df_preview[numeric_cols].describe() if numeric_cols else None,
                    "error": None,
                }
            except Exception as e:
                logging.warning("Failed to load sheet '%s' from %s: %s", sheet_name, path, e)
                sheet_data = {
                    "name": sheet_name,
                    "error": str(e),
                    "total_rows": 0,
                    "total_cols": 0,
                    "truncated": False,
                    "headers": [],
                    "rows": [],
                    "numeric_cols": [],
                    "null_counts": {},
                    "dtypes": {},
                    "describe": None,
                }

            result["sheets"].append(sheet_data)

    except ImportError as e:
        logging.error("Required library missing for Excel preview: %s", e)
        result["error"] = (
            f"Libreria mancante: {e}. Installa con: sudo apt install python3-openpyxl python3-pandas"
        )
    except MemoryError:
        logging.error("Out of memory reading %s", path)
        result["error"] = "File troppo grande: memoria insufficiente"
    except Exception as e:
        logging.warning("Unexpected error reading Excel file %s: %s", path, e)
        result["error"] = str(e)

    return result


# --------------------------------------------------------------------------- #
# Finestra
# --------------------------------------------------------------------------- #


class ExcelPreviewWindow(Gtk.Window):
    def __init__(self, path: str):
        filename = os.path.basename(path)
        super().__init__(title=f"{filename} — Excel Preview")
        self.set_default_size(WINDOW_W, WINDOW_H)
        self._path = path

        provider = Gtk.CssProvider()
        provider.load_from_data(CSS)
        Gtk.StyleContext.add_provider_for_display(
            self.get_display(), provider, Gtk.STYLE_PROVIDER_PRIORITY_APPLICATION
        )

        self._build_skeleton()
        threading.Thread(target=self._load, daemon=True).start()

    def _build_skeleton(self):
        self._root = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0)
        self.set_child(self._root)
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=12)
        box.set_valign(Gtk.Align.CENTER)
        box.set_halign(Gtk.Align.CENTER)
        box.set_vexpand(True)
        spinner = Gtk.Spinner()
        spinner.set_size_request(48, 48)
        spinner.start()
        lbl = Gtk.Label(label="Lettura file Excel…")
        lbl.add_css_class("dim-label")
        box.append(spinner)
        box.append(lbl)
        self._spinner_box = box
        self._root.append(box)

    def _load(self):
        data = read_excel(self._path)
        GLib.idle_add(self._on_loaded, data)

    def _on_loaded(self, data):
        self._root.remove(self._spinner_box)
        if data.get("error"):
            lbl = Gtk.Label(label=f"Errore: {data['error']}")
            lbl.set_margin_top(40)
            self._root.append(lbl)
            return False
        self._build_content(data)
        return False

    # ------------------------------------------------------------------ #
    # Contenuto principale
    # ------------------------------------------------------------------ #

    def _build_content(self, data: dict):
        sheets = data["sheets"]
        total_sheets = len(sheets)
        total_rows = sum(s.get("total_rows", 0) for s in sheets)

        # --- Info bar ---
        info_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=4)
        info_bar.add_css_class("xl-info-bar")
        info_bar.set_margin_start(4)
        info_bar.set_margin_end(8)

        self._stat(info_bar, "Fogli", str(total_sheets))
        self._stat(info_bar, "Righe tot.", f"{total_rows:,}")
        self._stat(info_bar, "Dimensione", data["file_size"])

        meta = data.get("metadata", {})
        if meta.get("Autore") and meta["Autore"] != "—":
            self._stat(info_bar, "Autore", meta["Autore"])
        if meta.get("Modificato") and meta["Modificato"] != "—":
            self._stat(info_bar, "Modificato", meta["Modificato"])

        self._root.append(info_bar)
        self._root.append(Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL))

        # --- Notebook principale: un tab per foglio ---
        if total_sheets == 1:
            # Un solo foglio: mostra direttamente senza tab foglio
            content = self._build_sheet_content(sheets[0])
            content.set_vexpand(True)
            self._root.append(content)
        else:
            sheets_nb = Gtk.Notebook()
            sheets_nb.set_vexpand(True)
            sheets_nb.set_tab_pos(Gtk.PositionType.LEFT)

            for sheet in sheets:
                tab_label = Gtk.Label(label=sheet["name"])
                tab_label.add_css_class("xl-sheet-tab")
                tab_label.set_max_width_chars(20)
                tab_label.set_ellipsize(Pango.EllipsizeMode.END)
                if sheet.get("error"):
                    safe_name = GLib.markup_escape_text(sheet["name"])
                    tab_label.set_markup(f"<span foreground='#d73a49'>{safe_name}</span>")

                content = self._build_sheet_content(sheet)
                sheets_nb.append_page(content, tab_label)

            self._root.append(sheets_nb)

        # --- Metadati tab sotto (se ci sono dati) ---
        if any(v != "—" for v in meta.values()):
            self._root.append(Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL))

        # --- Bottom bar ---
        bottom = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=8)
        bottom.set_margin_start(8)
        bottom.set_margin_end(8)
        bottom.set_margin_top(6)
        bottom.set_margin_bottom(6)

        path_lbl = Gtk.Label(label=self._path)
        path_lbl.add_css_class("dim-label")
        path_lbl.set_ellipsize(Pango.EllipsizeMode.START)
        path_lbl.set_hexpand(True)
        path_lbl.set_halign(Gtk.Align.START)
        bottom.append(path_lbl)

        open_btn = Gtk.Button(label="Apri con LibreOffice")
        open_btn.connect("clicked", self._open_editor)
        bottom.append(open_btn)

        self._root.append(Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL))
        self._root.append(bottom)

    def _open_editor(self, _btn):
        if not os.path.exists(self._path):
            logging.warning("xdg-open: file not found: %s", self._path)
            return
        try:
            subprocess.run(["xdg-open", self._path], check=False)
        except OSError as e:
            logging.warning("xdg-open failed for %s: %s", self._path, e)

    # ------------------------------------------------------------------ #
    # Contenuto di un singolo foglio
    # ------------------------------------------------------------------ #

    def _build_sheet_content(self, sheet: dict) -> Gtk.Widget:
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0)
        box.set_vexpand(True)

        if sheet.get("error"):
            lbl = Gtk.Label(label=f"Errore nel foglio '{sheet['name']}': {sheet['error']}")
            lbl.add_css_class("dim-label")
            lbl.set_margin_top(20)
            box.append(lbl)
            return box

        # Sub-info bar per questo foglio
        sub_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=12)
        sub_bar.set_margin_start(12)
        sub_bar.set_margin_end(12)
        sub_bar.set_margin_top(6)
        sub_bar.set_margin_bottom(6)

        info_lbl = Gtk.Label(
            label=f"{sheet['total_rows']:,} righe  ×  {sheet['total_cols']} colonne"
        )
        info_lbl.add_css_class("dim-label")
        info_lbl.set_halign(Gtk.Align.START)
        sub_bar.append(info_lbl)

        if sheet["truncated"]:
            trunc = Gtk.Label(label=f"⚠ Mostrate prime {PREVIEW_ROWS} righe")
            trunc.add_css_class("dim-label")
            trunc.set_halign(Gtk.Align.END)
            trunc.set_hexpand(True)
            sub_bar.append(trunc)

        box.append(sub_bar)
        box.append(Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL))

        # Notebook interno: Dati / Statistiche / Colonne
        inner_nb = Gtk.Notebook()
        inner_nb.set_vexpand(True)

        inner_nb.append_page(self._tab_data(sheet), Gtk.Label(label="📊 Dati"))

        if sheet.get("describe") is not None:
            inner_nb.append_page(self._tab_stats(sheet), Gtk.Label(label="📈 Statistiche"))

        inner_nb.append_page(self._tab_columns(sheet), Gtk.Label(label="🗂 Colonne"))

        box.append(inner_nb)
        return box

    # ------------------------------------------------------------------ #
    # Tab Dati
    # ------------------------------------------------------------------ #

    def _tab_data(self, sheet: dict) -> Gtk.Widget:
        scrolled = Gtk.ScrolledWindow()
        scrolled.set_vexpand(True)

        headers = sheet["headers"]
        rows = sheet["rows"]
        numeric_cols = sheet["numeric_cols"]
        dtypes = sheet["dtypes"]

        if not headers:
            lbl = Gtk.Label(label="Foglio vuoto.")
            lbl.add_css_class("dim-label")
            lbl.set_margin_top(20)
            scrolled.set_child(lbl)
            return scrolled

        col_types = [str] * len(headers)
        store = Gtk.ListStore(*col_types)
        for row in rows:
            # Assicura che la riga abbia il numero corretto di colonne
            padded = row + [""] * max(0, len(headers) - len(row))
            store.append(padded[: len(headers)])

        tv = Gtk.TreeView(model=store)
        tv.set_grid_lines(Gtk.TreeViewGridLines.BOTH)
        tv.add_css_class("mono")

        for i, h in enumerate(headers):
            is_numeric = h in [str(c) for c in numeric_cols]

            renderer = Gtk.CellRendererText()
            renderer.set_property("ellipsize", Pango.EllipsizeMode.END)
            if is_numeric:
                renderer.set_property("xalign", 1.0)
                renderer.set_property("foreground", "#0366d6")

            col = Gtk.TreeViewColumn()
            col.pack_start(renderer, True)
            col.add_attribute(renderer, "text", i)
            col.set_resizable(True)
            col.set_sizing(Gtk.TreeViewColumnSizing.FIXED)
            col.set_fixed_width(max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, len(str(h)) * 11 + 24)))
            col.set_sort_column_id(i)

            # Header con nome + tipo
            dtype_str = dtypes.get(h, "")
            header_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0)
            name_lbl = Gtk.Label()
            safe_h = GLib.markup_escape_text(str(h))
            name_lbl.set_markup(f"<b>{safe_h}</b>")
            type_lbl = Gtk.Label()
            color = "#0366d6" if is_numeric else "#6a737d"
            safe_dtype = GLib.markup_escape_text(dtype_str)
            type_lbl.set_markup(f"<span foreground='{color}' size='small'>{safe_dtype}</span>")
            header_box.append(name_lbl)
            header_box.append(type_lbl)
            header_box.show()
            col.set_widget(header_box)

            tv.append_column(col)

        scrolled.set_child(tv)
        return scrolled

    # ------------------------------------------------------------------ #
    # Tab Statistiche
    # ------------------------------------------------------------------ #

    def _tab_stats(self, sheet: dict) -> Gtk.Widget:
        scrolled = Gtk.ScrolledWindow()
        scrolled.set_vexpand(True)

        outer = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=8)
        outer.set_margin_start(16)
        outer.set_margin_end(16)
        outer.set_margin_top(12)
        outer.set_margin_bottom(12)

        desc = sheet["describe"]
        numeric_cols = [str(c) for c in sheet["numeric_cols"]]

        title = Gtk.Label()
        title.set_markup("<b>Statistiche descrittive (colonne numeriche)</b>")
        title.set_halign(Gtk.Align.START)
        outer.append(title)

        col_types = [str] * (len(numeric_cols) + 1)
        store = Gtk.ListStore(*col_types)
        for stat in desc.index:
            row = [stat] + [f"{desc.loc[stat, col]:.4g}" for col in numeric_cols]
            store.append(row)

        tv = Gtk.TreeView(model=store)
        tv.set_grid_lines(Gtk.TreeViewGridLines.HORIZONTAL)
        tv.add_css_class("mono")

        r = Gtk.CellRendererText()
        r.set_property("weight", Pango.Weight.BOLD)
        c = Gtk.TreeViewColumn("Statistica", r, text=0)
        c.set_min_width(100)
        tv.append_column(c)

        for i, col in enumerate(numeric_cols):
            r = Gtk.CellRendererText()
            r.set_property("xalign", 1.0)
            r.set_property("foreground", "#0366d6")
            c = Gtk.TreeViewColumn(col, r, text=i + 1)
            c.set_resizable(True)
            c.set_min_width(MIN_COL_WIDTH)
            tv.append_column(c)

        outer.append(tv)

        # Valori nulli
        null_counts = sheet.get("null_counts", {})
        nulls = {k: v for k, v in null_counts.items() if v > 0}

        null_title = Gtk.Label()
        null_title.set_markup("<b>Valori nulli per colonna</b>")
        null_title.set_halign(Gtk.Align.START)
        null_title.set_margin_top(12)
        outer.append(null_title)

        if nulls:
            total = sheet.get("total_rows", 1) or 1
            for col, count in nulls.items():
                pct = count / total * 100
                row_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=8)
                col_lbl = Gtk.Label(label=str(col))
                col_lbl.set_width_chars(24)
                col_lbl.set_halign(Gtk.Align.START)
                val_lbl = Gtk.Label(label=f"{count:,} ({pct:.1f}%)")
                val_lbl.add_css_class("dim-label")
                row_box.append(col_lbl)
                row_box.append(val_lbl)
                outer.append(row_box)
        else:
            ok = Gtk.Label(label="✓ Nessun valore nullo")
            ok.set_halign(Gtk.Align.START)
            outer.append(ok)

        scrolled.set_child(outer)
        return scrolled

    # ------------------------------------------------------------------ #
    # Tab Colonne
    # ------------------------------------------------------------------ #

    def _tab_columns(self, sheet: dict) -> Gtk.Widget:
        scrolled = Gtk.ScrolledWindow()
        scrolled.set_vexpand(True)

        headers = sheet["headers"]
        dtypes = sheet["dtypes"]
        null_counts = sheet.get("null_counts", {})
        numeric_cols = [str(c) for c in sheet["numeric_cols"]]
        total_rows = sheet.get("total_rows", 0) or 1

        store = Gtk.ListStore(str, str, str, str, str)
        for i, h in enumerate(headers):
            h_str = str(h)
            dtype = dtypes.get(h_str, "?")
            nulls = null_counts.get(h_str, 0)
            null_str = f"{nulls:,} ({nulls / total_rows * 100:.1f}%)" if nulls > 0 else "—"
            kind = "numerico" if h_str in numeric_cols else "testo"
            store.append([str(i + 1), h_str, dtype, kind, null_str])

        tv = Gtk.TreeView(model=store)
        tv.set_grid_lines(Gtk.TreeViewGridLines.HORIZONTAL)

        defs = [
            ("#", 40),
            ("Nome colonna", 220),
            ("Tipo", 120),
            ("Categoria", 100),
            ("Valori nulli", 140),
        ]
        for i, (label, width) in enumerate(defs):
            r = Gtk.CellRendererText()
            if i == 3:  # categoria: colore diverso per numerico
                c = Gtk.TreeViewColumn(label)
                c.pack_start(r, True)
                c.set_min_width(width)

                def set_kind_color(col, cell, model, it, _):
                    val = model.get_value(it, 3)
                    cell.set_property("text", val)
                    cell.set_property("foreground", "#0366d6" if val == "numerico" else "#6a737d")

                c.set_cell_data_func(r, set_kind_color, None)
            else:
                c = Gtk.TreeViewColumn(label, r, text=i)
                c.set_resizable(True)
                c.set_min_width(width)
            tv.append_column(c)

        scrolled.set_child(tv)
        return scrolled

    # ------------------------------------------------------------------ #
    # Helpers
    # ------------------------------------------------------------------ #

    def _stat(self, box, title, value):
        sb = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=2)
        sb.add_css_class("xl-stat-box")
        t = Gtk.Label(label=title)
        t.add_css_class("xl-stat-title")
        t.set_halign(Gtk.Align.START)
        v = Gtk.Label(label=str(value))
        v.add_css_class("xl-stat-value")
        v.set_halign(Gtk.Align.START)
        sb.append(t)
        sb.append(v)
        box.append(sb)


# --------------------------------------------------------------------------- #
# Extension
# --------------------------------------------------------------------------- #


class ExcelPreviewExtension(GObject.GObject, Nautilus.MenuProvider):
    def get_file_items(self, files):
        if len(files) != 1:
            return []
        f = files[0]
        ext = os.path.splitext(f.get_name().lower())[1]
        if ext not in SUPPORTED_EXTENSIONS:
            return []

        item = Nautilus.MenuItem(
            name="ExcelPreview::show",
            label="Anteprima Excel",
            tip=f"Mostra anteprima di {f.get_name()}",
        )
        item.connect("activate", self._on_activate, f.get_location().get_path())
        return [item]

    def get_background_items(self, folder):
        return []

    def _on_activate(self, _item, path):
        win = ExcelPreviewWindow(path)
        win.present()
