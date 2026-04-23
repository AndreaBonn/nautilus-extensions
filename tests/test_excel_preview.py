"""Tests for excel-preview pure functions."""

import os
import tempfile
from pathlib import Path

import openpyxl


def _load_functions():
    source = (Path(__file__).parent.parent / "excel-preview" / "excel_preview.py").read_text()
    namespace = {}
    exec("import os, logging, threading", namespace)
    lines = source.split("\n")
    safe_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith(("import gi", "from gi.", "gi.require_version")):
            continue
        if stripped.startswith("class ") and ("Gtk." in stripped or "GObject." in stripped):
            break
        safe_lines.append(line)
    exec("\n".join(safe_lines), namespace)
    return namespace


_ns = _load_functions()
fmt_size = _ns["fmt_size"]
read_excel = _ns["read_excel"]
PREVIEW_ROWS = _ns["PREVIEW_ROWS"]


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #


def _make_xlsx(rows: list[list], sheet_name: str = "Sheet1") -> str:
    """Create a temporary .xlsx file with given rows (first row = headers)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _make_xlsx_multi(sheets: dict[str, list[list]]) -> str:
    """Create a temporary .xlsx with multiple named sheets."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# --------------------------------------------------------------------------- #
# TestFmtSize
# --------------------------------------------------------------------------- #


class TestFmtSize:
    def test_fmt_size_bytes_returns_b_unit(self):
        assert fmt_size(0) == "0.0 B"

    def test_fmt_size_999_bytes(self):
        assert fmt_size(999) == "999.0 B"

    def test_fmt_size_exactly_1kb(self):
        assert fmt_size(1024) == "1.0 KB"

    def test_fmt_size_exactly_1mb(self):
        assert fmt_size(1024 * 1024) == "1.0 MB"

    def test_fmt_size_exactly_1gb(self):
        assert fmt_size(1024**3) == "1.0 GB"

    def test_fmt_size_large_terabytes(self):
        assert fmt_size(1024**4) == "1.0 TB"

    def test_fmt_size_2kb(self):
        assert fmt_size(2048) == "2.0 KB"

    def test_fmt_size_1_5mb(self):
        result = fmt_size(int(1.5 * 1024 * 1024))
        assert "MB" in result
        assert "1.5" in result


# --------------------------------------------------------------------------- #
# TestReadExcel — basic structure
# --------------------------------------------------------------------------- #


class TestReadExcelBasicStructure:
    def test_read_excel_returns_dict_with_required_keys(self):
        path = _make_xlsx([["name", "age"], ["Alice", 30]])
        try:
            result = read_excel(path)
            assert "path" in result
            assert "error" in result
            assert "sheets" in result
            assert "metadata" in result
            assert "file_size" in result
        finally:
            os.unlink(path)

    def test_read_excel_path_field_matches_input(self):
        path = _make_xlsx([["col"], ["val"]])
        try:
            result = read_excel(path)
            assert result["path"] == path
        finally:
            os.unlink(path)

    def test_read_excel_no_error_on_valid_file(self):
        path = _make_xlsx([["a", "b"], [1, 2]])
        try:
            result = read_excel(path)
            assert result["error"] is None
        finally:
            os.unlink(path)

    def test_read_excel_file_size_is_non_empty_string(self):
        path = _make_xlsx([["x"], [1]])
        try:
            result = read_excel(path)
            assert isinstance(result["file_size"], str)
            assert len(result["file_size"]) > 0
        finally:
            os.unlink(path)

    def test_read_excel_metadata_contains_fogli_key(self):
        path = _make_xlsx([["col"], [1]])
        try:
            result = read_excel(path)
            assert "Fogli" in result["metadata"]
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — single sheet content
# --------------------------------------------------------------------------- #


class TestReadExcelSingleSheet:
    def test_read_excel_single_sheet_returns_one_sheet(self):
        path = _make_xlsx([["name", "score"], ["Alice", 95], ["Bob", 80]])
        try:
            result = read_excel(path)
            assert len(result["sheets"]) == 1
        finally:
            os.unlink(path)

    def test_read_excel_headers_match_first_row(self):
        path = _make_xlsx([["name", "score"], ["Alice", 95]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["headers"] == ["name", "score"]
        finally:
            os.unlink(path)

    def test_read_excel_rows_contain_data_values(self):
        path = _make_xlsx([["name", "score"], ["Alice", 95], ["Bob", 80]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            # rows are stringified by astype(str)
            assert len(sheet["rows"]) == 2
            assert sheet["rows"][0][0] == "Alice"
        finally:
            os.unlink(path)

    def test_read_excel_total_cols_matches_header_count(self):
        path = _make_xlsx([["a", "b", "c"], [1, 2, 3]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["total_cols"] == 3
        finally:
            os.unlink(path)

    def test_read_excel_sheet_name_preserved(self):
        path = _make_xlsx([["x"], [1]], sheet_name="MySheet")
        try:
            result = read_excel(path)
            assert result["sheets"][0]["name"] == "MySheet"
        finally:
            os.unlink(path)

    def test_read_excel_sheet_error_field_is_none_on_valid_sheet(self):
        path = _make_xlsx([["col"], [1]])
        try:
            result = read_excel(path)
            assert result["sheets"][0]["error"] is None
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — multi-sheet
# --------------------------------------------------------------------------- #


class TestReadExcelMultiSheet:
    def test_read_excel_multi_sheet_returns_all_sheets(self):
        path = _make_xlsx_multi(
            {
                "Alpha": [["a"], [1]],
                "Beta": [["b"], [2]],
                "Gamma": [["c"], [3]],
            }
        )
        try:
            result = read_excel(path)
            assert len(result["sheets"]) == 3
        finally:
            os.unlink(path)

    def test_read_excel_multi_sheet_names_are_correct(self):
        path = _make_xlsx_multi(
            {
                "Sales": [["revenue"], [100]],
                "Costs": [["expense"], [50]],
            }
        )
        try:
            result = read_excel(path)
            names = [s["name"] for s in result["sheets"]]
            assert "Sales" in names
            assert "Costs" in names
        finally:
            os.unlink(path)

    def test_read_excel_multi_sheet_metadata_fogli_count_is_correct(self):
        path = _make_xlsx_multi(
            {
                "A": [["x"], [1]],
                "B": [["y"], [2]],
            }
        )
        try:
            result = read_excel(path)
            assert result["metadata"]["Fogli"] == "2"
        finally:
            os.unlink(path)

    def test_read_excel_multi_sheet_each_has_independent_headers(self):
        path = _make_xlsx_multi(
            {
                "People": [["name", "age"], ["Alice", 30]],
                "Products": [["sku", "price", "stock"], ["X1", 9.99, 50]],
            }
        )
        try:
            result = read_excel(path)
            sheets_by_name = {s["name"]: s for s in result["sheets"]}
            assert sheets_by_name["People"]["headers"] == ["name", "age"]
            assert sheets_by_name["Products"]["headers"] == ["sku", "price", "stock"]
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — empty sheet
# --------------------------------------------------------------------------- #


class TestReadExcelEmptySheet:
    def test_read_excel_empty_sheet_has_empty_headers(self):
        path = _make_xlsx([])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["headers"] == []
        finally:
            os.unlink(path)

    def test_read_excel_empty_sheet_has_empty_rows(self):
        path = _make_xlsx([])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["rows"] == []
        finally:
            os.unlink(path)

    def test_read_excel_empty_sheet_total_cols_is_zero(self):
        path = _make_xlsx([])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["total_cols"] == 0
        finally:
            os.unlink(path)

    def test_read_excel_header_only_sheet_has_no_rows(self):
        path = _make_xlsx([["col_a", "col_b"]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["headers"] == ["col_a", "col_b"]
            assert sheet["rows"] == []
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — numeric columns
# --------------------------------------------------------------------------- #


class TestReadExcelNumericColumns:
    def test_read_excel_numeric_cols_identifies_integer_column(self):
        path = _make_xlsx([["name", "score"], ["Alice", 95], ["Bob", 80]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert "score" in sheet["numeric_cols"]
        finally:
            os.unlink(path)

    def test_read_excel_numeric_cols_excludes_string_column(self):
        path = _make_xlsx([["name", "score"], ["Alice", 95], ["Bob", 80]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert "name" not in sheet["numeric_cols"]
        finally:
            os.unlink(path)

    def test_read_excel_numeric_cols_identifies_float_column(self):
        path = _make_xlsx([["item", "price"], ["apple", 1.5], ["bread", 2.3]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert "price" in sheet["numeric_cols"]
        finally:
            os.unlink(path)

    def test_read_excel_dtypes_dict_has_entry_for_each_header(self):
        path = _make_xlsx([["a", "b", "c"], [1, "x", 3.0]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            for header in sheet["headers"]:
                assert header in sheet["dtypes"]
        finally:
            os.unlink(path)

    def test_read_excel_no_numeric_cols_when_all_text(self):
        path = _make_xlsx([["first", "last"], ["Alice", "Smith"], ["Bob", "Jones"]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["numeric_cols"] == []
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — null value counting
# --------------------------------------------------------------------------- #


class TestReadExcelNullCounting:
    def test_read_excel_null_counts_zero_when_no_nulls(self):
        path = _make_xlsx([["a", "b"], [1, 2], [3, 4]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["null_counts"]["a"] == 0
            assert sheet["null_counts"]["b"] == 0
        finally:
            os.unlink(path)

    def test_read_excel_null_counts_detects_missing_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["Alice", 10])
        ws.append(["Bob", None])  # explicit None → NaN in pandas
        ws.append(["Carol", 30])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        try:
            result = read_excel(tmp.name)
            sheet = result["sheets"][0]
            assert sheet["null_counts"]["value"] == 1
        finally:
            os.unlink(tmp.name)

    def test_read_excel_null_counts_keys_match_headers(self):
        path = _make_xlsx([["x", "y", "z"], [1, 2, 3]])
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            for header in sheet["headers"]:
                assert header in sheet["null_counts"]
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — truncation behavior
# --------------------------------------------------------------------------- #


class TestReadExcelTruncation:
    def test_read_excel_truncated_false_when_rows_within_limit(self):
        rows = [["idx"]] + [[i] for i in range(10)]
        path = _make_xlsx(rows)
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["truncated"] is False
        finally:
            os.unlink(path)

    def test_read_excel_truncated_true_when_rows_exceed_preview_limit(self):
        # PREVIEW_ROWS + 5 data rows → truncated
        rows = [["idx"]] + [[i] for i in range(PREVIEW_ROWS + 5)]
        path = _make_xlsx(rows)
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["truncated"] is True
        finally:
            os.unlink(path)

    def test_read_excel_preview_rows_capped_at_preview_limit(self):
        rows = [["idx"]] + [[i] for i in range(PREVIEW_ROWS + 10)]
        path = _make_xlsx(rows)
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert len(sheet["rows"]) <= PREVIEW_ROWS
        finally:
            os.unlink(path)

    def test_read_excel_total_rows_reflects_full_count_not_preview(self):
        rows = [["idx"]] + [[i] for i in range(PREVIEW_ROWS + 20)]
        path = _make_xlsx(rows)
        try:
            result = read_excel(path)
            sheet = result["sheets"][0]
            assert sheet["total_rows"] > PREVIEW_ROWS
        finally:
            os.unlink(path)


# --------------------------------------------------------------------------- #
# TestReadExcel — error handling
# --------------------------------------------------------------------------- #


class TestReadExcelErrorHandling:
    def test_read_excel_nonexistent_file_sets_error_field(self):
        result = read_excel("/nonexistent/path/file.xlsx")
        assert result["error"] is not None
        assert isinstance(result["error"], str)

    def test_read_excel_nonexistent_file_returns_empty_sheets(self):
        result = read_excel("/nonexistent/path/file.xlsx")
        assert result["sheets"] == []

    def test_read_excel_nonexistent_file_path_field_preserved(self):
        path = "/nonexistent/path/file.xlsx"
        result = read_excel(path)
        assert result["path"] == path

    def test_read_excel_corrupted_file_sets_error_field(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(b"this is not a valid xlsx file at all!!!")
        tmp.close()
        try:
            result = read_excel(tmp.name)
            assert result["error"] is not None
            assert isinstance(result["error"], str)
        finally:
            os.unlink(tmp.name)

    def test_read_excel_corrupted_file_returns_empty_sheets(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(b"\x00\x01\x02\x03 garbage bytes not xlsx")
        tmp.close()
        try:
            result = read_excel(tmp.name)
            assert result["sheets"] == []
        finally:
            os.unlink(tmp.name)
